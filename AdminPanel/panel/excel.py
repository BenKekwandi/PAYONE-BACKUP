from openpyxl import Workbook
import os
from django.conf import settings
from django.http import HttpResponse, Http404
from django.http import FileResponse, JsonResponse
from .models import TransactionModel, CompanyModel, SalespersonModel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime


def getTransactions():
    data = list(TransactionModel.objects.all().values())
    allTransactions = []
    for trx in data:
        salesperson = SalespersonModel.objects.get(
            id=trx['salesperson_id'])
        company = CompanyModel.objects.get(id=salesperson.company_id)
        trx['salesperson'] = salesperson.name+' '+salesperson.surname
        trx['company'] = company.company_name
        trx['transaction_date'] = datetime.fromtimestamp(
            float(trx['start_timestamp']) / 1000).date()
        allTransactions.append(trx)
    if len(allTransactions) == 0:
        return None
    else:
        return allTransactions


def getReportExcelWorkbook():
    data = getTransactions()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DataSheet"
    headers = list(data[0].keys())
    sheet.append(headers)
    for item in data:
        sheet.append(list(item.values()))

    column_width = 45
    for column_idx in range(1, len(headers) + 1):
        sheet.column_dimensions[sheet.cell(
            row=1, column=column_idx).column_letter].width = column_width

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(
        style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border

    header_fill = PatternFill(start_color='FFFF00',
                              end_color='FFFF00', fill_type='solid')
    for cell in sheet[1]:
        cell.fill = header_fill
    return workbook


def getFullBankReport(request):
    data = list(TransactionModel.objects.all().values())
    allTransactions = []
    for trx in data:
        salesperson = SalespersonModel.objects.get(
            id=trx['salesperson_id'])
        company = CompanyModel.objects.get(id=salesperson.company_id)
        trx['salesperson'] = salesperson.name+' '+salesperson.surname
        trx['company'] = company.company_name
        trx['transaction_date'] = datetime.fromtimestamp(
            float(trx['start_timestamp']) / 1000).date()
        allTransactions.append(trx)
    if len(allTransactions) == 0:
        return HttpResponse(f"No transaction recorded yet")
    return getReportFile(allTransactions)


def getMonthlyBankReport(request):
    data = list(TransactionModel.objects.all().values())
    allTransactions = []
    today = datetime.now().date()
    for trx in data:
        salesperson = SalespersonModel.objects.get(
            id=trx['salesperson_id'])
        company = CompanyModel.objects.get(id=salesperson.company_id)
        trx['Salesperson'] = salesperson.name+' '+salesperson.surname
        trx['Company'] = company.company_name
        trx['Transaction Date'] = datetime.fromtimestamp(
            float(trx['start_timestamp']) / 1000).date()
        transaction_date = trx['Transaction Date']
        if transaction_date.month == today.month:
            allTransactions.append(trx)
    if len(allTransactions) == 0:
        return HttpResponse(f"No data recorded on {datetime.now()}")
    else:
        return getReportFile(allTransactions)


def getDailyBankReport(request):
    data = list(TransactionModel.objects.all().values())
    dailyTransactions = []
    today = datetime.now().date()

    for trx in data:
        salesperson = SalespersonModel.objects.get(
            id=trx['salesperson_id'])
        company = CompanyModel.objects.get(id=salesperson.company_id)
        trx['salesperson'] = salesperson.name + ' ' + salesperson.surname
        trx['company'] = company.company_name
        transaction_date = datetime.fromtimestamp(
            float(trx['start_timestamp']) / 1000).date()
        if transaction_date == today:
            dailyTransactions.append(trx)
    if len(dailyTransactions) == 0:
        return HttpResponse(f"No data recorded on {datetime.now()}")
    else:
        return getReportFile(dailyTransactions)


def getReportFile(data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DataSheet"
    headers = list(data[0].keys())
    sheet.append(headers)
    for item in data:
        sheet.append(list(item.values()))

    column_width = 45
    for column_idx in range(1, len(headers) + 1):
        sheet.column_dimensions[sheet.cell(
            row=1, column=column_idx).column_letter].width = column_width

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(
        style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border

    header_fill = PatternFill(start_color='FFFF00',
                              end_color='FFFF00', fill_type='solid')
    for cell in sheet[1]:
        cell.fill = header_fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename={datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.xlsx"
    workbook.save(response)
    return response


def getWorkbook(data):
    columnSize = [13,30,18,13,40]
    workbook = Workbook()
    sheet = workbook.active
    headers = list(data[0].keys())
    sheet.append(headers)
    for item in data:
        sheet.append(list(item.values()))
    column_width = 30
    for column_idx in range(1, len(headers) + 1):
        sheet.column_dimensions[sheet.cell(
            row=1, column=column_idx).column_letter].width = columnSize[column_idx-1]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(
        style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
    table = Table(displayName="DataTable", ref=sheet.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)
    
    return workbook